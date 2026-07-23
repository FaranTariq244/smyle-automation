"""Parse a TikTok To-Ship sync run log and build an Excel summary.

Columns: TikTok Order # | Shopify Order # | Fulfilment Link | Tracking # |
         Status (what we did).
"""
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

LOG = Path(sys.argv[1])
OUT = Path(sys.argv[2])

text = LOG.read_text(encoding="utf-8", errors="replace")
# Strip the "2026-07-23 14:33:55,560 [INFO] " prefix from every line.
lines = [re.sub(r"^\d{4}-\d\d-\d\d \d\d:\d\d:\d\d,\d+ \[\w+\] ?", "", ln)
         for ln in text.splitlines()]

# Split the FINAL SUMMARY into per-order blocks separated by dashed lines.
blocks, cur = [], []
in_summary = False
for ln in lines:
    if "FINAL SUMMARY" in ln:
        in_summary = True
        continue
    if not in_summary:
        continue
    if ln.startswith("TikTok order :"):
        if cur:
            blocks.append(cur)
        cur = [ln]
    elif cur is not None and cur:
        if set(ln.strip()) <= {"-"} and ln.strip():
            continue  # dashed separator
        if set(ln.strip()) <= {"="} and ln.strip():
            break      # end of summary section
        cur.append(ln)
if cur:
    blocks.append(cur)

rows = []
for b in blocks:
    txt = "\n".join(b)
    tiktok = re.search(r"TikTok order : (\S+)", txt)
    shopify = re.search(r"Shopify order: (\S+)", txt)
    detail = re.search(r"Detail page  : (\S+)", txt)
    tracking = re.search(r"Tracking     : (\S+)", txt)
    ref = re.search(r"Nic\.Oud ref (\d+) \| status (\w+)", txt)

    tiktok_id = tiktok.group(1) if tiktok else ""
    shopify_no = shopify.group(1) if shopify else ""
    link = detail.group(1) if detail else ""
    track_no = tracking.group(1) if tracking else ""

    if "NOT FOUND for reference" in txt:
        status = "Skipped - not found in my-fulfilment.com (not synced there yet)"
    elif tracking:
        status = f"UPLOADED to TikTok (tracking submitted, ref {ref.group(1)}, status {ref.group(2)})"
    elif "NOT SHIPPED YET" in txt:
        st = ref.group(2) if ref else "?"
        status = f"Skipped - not shipped yet (fulfilment status: {st})"
    else:
        status = "Skipped - no tracking"

    rows.append([tiktok_id, shopify_no, link, track_no, status])

# Add TikTok orders that never matched in Shopify (logged in Phase 2).
m = re.search(r"Not in Shopify \(likely not synced yet\): (.+)", "\n".join(lines))
if m:
    for oid in [x.strip() for x in m.group(1).split(",") if x.strip()]:
        rows.append([oid, "NOT FOUND",  "", "",
                     "Skipped - not synced to Shopify yet (no Shopify order to match)"])

# Sort: uploaded first, then not-shipped, then not-found.
def sort_key(r):
    if r[4].startswith("UPLOADED"):
        return 0
    if r[4].startswith("Skipped - not shipped"):
        return 1
    if "not found" in r[4]:
        return 2
    return 3
rows.sort(key=sort_key)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "TikTok To-Ship 2026-07-23"

headers = ["TikTok Order #", "Shopify Order #", "Fulfilment Link",
           "Tracking #", "Status (what we did)"]
ws.append(headers)

hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

green = PatternFill("solid", fgColor="E2EFDA")
yellow = PatternFill("solid", fgColor="FFF2CC")
red = PatternFill("solid", fgColor="FCE4D6")
for r, data in enumerate(rows, start=2):
    ws.append(data)
    if data[4].startswith("UPLOADED"):
        fill = green
    elif data[4].startswith("Skipped - not shipped"):
        fill = yellow
    else:
        fill = red
    for c in range(1, 6):
        cell = ws.cell(row=r, column=c)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(c == 5))
    # Make the fulfilment link clickable.
    if data[2]:
        link_cell = ws.cell(row=r, column=3)
        link_cell.hyperlink = data[2]
        link_cell.font = Font(color="0563C1", underline="single")

widths = [22, 18, 60, 18, 55]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:E{len(rows) + 1}"

wb.save(OUT)

# Retention: keep only the newest MAX_REPORTS Excel reports in this folder;
# delete older ones so the directory doesn't grow without bound.
MAX_REPORTS = 20
report_files = sorted(
    OUT.parent.glob("TikTok_ToShip_Report_*.xlsx"),
    key=lambda p: p.stat().st_mtime,
    reverse=True,
)
for old in report_files[MAX_REPORTS:]:
    try:
        old.unlink()
        print(f"Pruned old report (keeping newest {MAX_REPORTS}): {old.name}")
    except OSError as exc:
        print(f"Could not prune {old.name}: {exc}")

uploaded = sum(1 for r in rows if r[4].startswith("UPLOADED"))
not_ship = sum(1 for r in rows if r[4].startswith("Skipped - not shipped"))
not_found = sum(1 for r in rows if "not found" in r[4])
not_synced = sum(1 for r in rows if "not synced to Shopify" in r[4])
# SUMMARY line — the supervisor reads this for the Slack report. Keep the
# format stable: "SUMMARY: processed=.. uploaded=.. pending=.. unmatched=.."
processed = len(rows)
unmatched = not_found + not_synced
print(f"SUMMARY: processed={processed} uploaded={uploaded} "
      f"pending={not_ship} unmatched={unmatched} "
      f"(pending=not shipped yet, unmatched={not_found} not in fulfilment + "
      f"{not_synced} not in Shopify)")
print(f"Wrote {OUT} with {processed} orders: {uploaded} uploaded, "
      f"{not_ship} not shipped yet, {not_found} not in fulfilment, "
      f"{not_synced} not in Shopify.")
