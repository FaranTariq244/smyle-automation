"""
RAW proof-of-concept: TikTok Seller Center "To ship" order export.

Workflow (first candidate for the new "Workflows" dashboard section):
    1. Open seller-nl.tiktok.com Manage orders > To ship
    2. Click Export
    3. In the Export orders dialog: select "All orders awaiting shipping",
       format CSV, click Export
    4. Poll Export History until the new report appears (TikTok takes ~2 min)
    5. Click Download on the new report
    6. Read the downloaded CSV and print the extracted data to the log

Then, for each order that has tracking:
    7. On the To ship list, click "Add tracking info" on that order's row
    8. If the "Combine orders" dialog appears, click "Combine orders and
       continue" (it is skipped when there is nothing to combine)
    9. Fill Tracking ID + Shipping provider, leave Receipt ID blank, and
       click "Submit N parcel(s)"

Fulfilment lookups go over HTTP (services/fulfilment/client.py); the browser
is only needed for TikTok Seller Center itself.

Uses the same persistent chrome_profile as the other extractors. First run
requires a one-time manual login to TikTok Seller Center in the opened
browser window; the session persists afterwards.

Usage:
    python tiktok_toship_export_raw.py              # visible browser
    python tiktok_toship_export_raw.py --headless   # after login is saved

    # verify the Add tracking info flow without submitting anything:
    python tiktok_toship_export_raw.py --tracking-test <order_id> <tracking> [carrier]
    #   ... and add --submit to actually send it

Environment toggles:
    TIKTOK_UPLOAD_DRY_RUN=1         fill the tracking page but never Submit
    TIKTOK_UPLOAD_USE_TEMPLATE=1    deprecated xlsx template upload path
    TIKTOK_FULFILMENT_USE_BROWSER=1 deprecated Selenium fulfilment lookups
"""

import csv
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path

from selenium.common.exceptions import WebDriverException
from selenium.webdriver.common.by import By

from browser_manager import BrowserManager

PROJECT_ROOT = Path(__file__).resolve().parent
DOWNLOAD_DIR = PROJECT_ROOT / "reports" / "tiktok_exports"
LOG_DIR = PROJECT_ROOT / "logs"

ORDERS_URL = "https://seller-nl.tiktok.com/order?order_status[]=1&selected_sort=1&tab=to_ship"
FULFILMENT_ORDERS_URL = "https://www.my-fulfilment.com/orders/"
SHIP_UPLOAD_URL = "https://seller-nl.tiktok.com/order/seller-shipping/upload?shop_region=NL"

# Fulfilment "Shipper" -> TikTok template "Shipping provider name" (must
# exactly match the template's shipping_provider_name_drop_lis sheet).
CARRIER_MAP = {
    "POSTNL": "PostNL Netherlands",
    "GLS": "GLS Netherlands",
    "DHL": "DHL_Netherlands",
    "DPD": "DPD Netherlands",
    "UPS": "UPS Netherlands",
    "PEDDLER": "Peddler Netherlands",
}


def tiktok_provider(shipper):
    """Map a my-fulfilment.com "Shipper" (e.g. 'PostNL') to TikTok's provider
    name (e.g. 'PostNL Netherlands'). Matches on the leading word."""
    key = (shipper or "").strip().upper().split()[0] if shipper else ""
    return CARRIER_MAP.get(key, shipper)

LOGIN_TIMEOUT = 600          # max seconds to wait for manual login
EXPORT_READY_TIMEOUT = 360   # max seconds to wait for TikTok to generate the report
DOWNLOAD_TIMEOUT = 120       # max seconds to wait for the file to land on disk

log = logging.getLogger("tiktok_toship")


def setup_logging():
    LOG_DIR.mkdir(exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(LOG_DIR / "tiktok_toship_export.log", encoding="utf-8"),
        ],
    )


def log_phase(title):
    """Print a clear, scannable phase banner to the log."""
    log.info("")
    log.info("==================== %s ====================", title)


def screenshot(driver, name):
    path = LOG_DIR / f"tiktok_{name}_{datetime.now():%H%M%S}.png"
    try:
        driver.save_screenshot(str(path))
        log.info("Screenshot saved: %s", path)
    except WebDriverException:
        pass


def find_visible(driver, xpath):
    """Return the first displayed element matching xpath, else None."""
    for el in driver.find_elements(By.XPATH, xpath):
        try:
            if el.is_displayed():
                return el
        except WebDriverException:
            continue
    return None


def wait_for(driver, xpath, timeout, what):
    log.info("Waiting for %s ...", what)
    deadline = time.time() + timeout
    while time.time() < deadline:
        el = find_visible(driver, xpath)
        if el:
            return el
        time.sleep(2)
    screenshot(driver, f"timeout_{what.replace(' ', '_')}")
    raise TimeoutError(f"Timed out waiting for {what}")


def js_click(driver, el):
    """Click via JS — immune to overlay/scroll interception in the drawer."""
    driver.execute_script("arguments[0].scrollIntoView({block:'center'}); arguments[0].click();", el)


def ensure_logged_in(driver, headless):
    driver.get(ORDERS_URL)
    time.sleep(10)
    if "/account/login" not in driver.current_url:
        log.info("Already logged in to Seller Center.")
        return
    if headless:
        raise RuntimeError(
            "Not logged in to TikTok Seller Center and running headless — "
            "run once WITHOUT --headless to log in manually; the session is "
            "then saved to chrome_profile and headless runs will work."
        )
    log.warning("=" * 60)
    log.warning("MANUAL LOGIN REQUIRED — please log in to TikTok Seller")
    log.warning("Center in the opened Chrome window (waiting up to %d min).", LOGIN_TIMEOUT // 60)
    log.warning("=" * 60)
    deadline = time.time() + LOGIN_TIMEOUT
    while time.time() < deadline:
        if "/order" in driver.current_url and "/account/login" not in driver.current_url:
            log.info("Login detected, session saved to chrome_profile.")
            time.sleep(5)
            return
        time.sleep(3)
    raise TimeoutError("Manual login was not completed in time.")


def first_history_entry(driver):
    """Return the report name of the top Export History row (or None).

    The heading TikTok renders is 'Export history (2)' — lowercase 'h' and a
    trailing count — so the match must be case-insensitive. We also anchor on
    the heading's own text() node (not normalize-space(.), which matches the
    whole dialog container as an ancestor); the report rows are only in the
    following:: axis of the *leaf* heading element, not of the container.
    """
    el = find_visible(
        driver,
        "//*[contains("
        "translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),"
        " 'export history')]"
        "/following::*[contains(normalize-space(text()), '.csv') "
        "or contains(normalize-space(text()), '.xlsx')][1]",
    )
    return el.text.strip() if el else None


def run_export(driver):
    """Steps 2-3: open the Export dialog, pick options, start the export.

    Returns the report name that was at the top of Export History before
    the new export, so the caller can detect the new entry.
    """
    export_btn = wait_for(
        driver,
        "//button[.//text()[normalize-space()='Export']]",
        60,
        "Export button on orders page",
    )
    js_click(driver, export_btn)
    log.info("Clicked Export — waiting for the Export orders dialog.")

    wait_for(
        driver,
        "//*[contains(normalize-space(text()), 'Export orders')]",
        30,
        "Export orders dialog",
    )
    time.sleep(2)
    previous_top = first_history_entry(driver)
    log.info("Top of Export History before export: %s", previous_top)

    awaiting = wait_for(
        driver,
        "//*[contains(normalize-space(text()), 'All orders awaiting shipping')]",
        20,
        "'All orders awaiting shipping' option",
    )
    js_click(driver, awaiting)
    log.info("Selected: All orders awaiting shipping")

    csv_opt = wait_for(
        driver,
        "//*[normalize-space(text())='CSV']",
        20,
        "CSV format option",
    )
    js_click(driver, csv_opt)
    log.info("Selected format: CSV")
    time.sleep(1)

    # The dialog's Export button is the last visible one (the page header
    # also has an Export button behind the drawer overlay).
    buttons = [
        el for el in driver.find_elements(
            By.XPATH, "//button[.//text()[normalize-space()='Export']]")
        if el.is_displayed()
    ]
    if not buttons:
        screenshot(driver, "no_dialog_export_btn")
        raise RuntimeError("Export button inside the dialog not found")
    js_click(driver, buttons[-1])
    log.info("Clicked Export in the dialog — TikTok is generating the report.")
    return previous_top


def wait_for_new_report(driver, previous_top):
    """Step 4: poll Export History until a new entry appears at the top."""
    deadline = time.time() + EXPORT_READY_TIMEOUT
    while time.time() < deadline:
        time.sleep(10)
        top = first_history_entry(driver)
        log.info("Export History top entry: %s", top)
        if top and top != previous_top:
            log.info("New report detected: %s", top)
            return top
    screenshot(driver, "export_never_appeared")
    raise TimeoutError("New export never appeared in Export History")


def download_report(driver, report_name):
    """Step 5: click Download on the row of the given report; return the file path."""
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    driver.execute_cdp_cmd(
        "Browser.setDownloadBehavior",
        {"behavior": "allow", "downloadPath": str(DOWNLOAD_DIR)},
    )
    before = {p.name for p in DOWNLOAD_DIR.glob("*")}

    # The Download button only appears in the report's row once TikTok
    # finishes generating the file. Climb the DOM from the report-name cell
    # to its own row and accept only a button inside that row — a sibling
    # row's button (e.g. the previous, already-finished report) must never
    # match, or we'd silently download stale data.
    find_row_button_js = """
        const name = arguments[0];
        const leaves = [...document.querySelectorAll('*')]
            .filter(e => e.children.length === 0 && e.textContent.trim() === name);
        for (const leaf of leaves) {
            let node = leaf;
            while (node && node !== document.body) {
                const btns = [...node.querySelectorAll('button')]
                    .filter(b => b.textContent.includes('Download'));
                const reports = (node.textContent.match(/\\.csv|\\.xlsx/g) || []).length;
                if (reports > 1 || btns.length > 1) break;  // climbed past the row
                if (btns.length === 1) return btns[0];
                node = node.parentElement;
            }
        }
        return null;
    """
    log.info("Waiting for Download button in the row of '%s' ...", report_name)
    deadline = time.time() + EXPORT_READY_TIMEOUT
    download_btn = None
    while time.time() < deadline:
        download_btn = driver.execute_script(find_row_button_js, report_name)
        if download_btn:
            break
        time.sleep(5)
    if not download_btn:
        screenshot(driver, "no_row_download_btn")
        raise TimeoutError(f"Download button for '{report_name}' never appeared")
    js_click(driver, download_btn)
    log.info("Clicked Download — waiting for the file.")

    deadline = time.time() + DOWNLOAD_TIMEOUT
    while time.time() < deadline:
        new_files = [
            p for p in DOWNLOAD_DIR.glob("*")
            if p.name not in before and not p.name.endswith(".crdownload")
        ]
        if new_files and not list(DOWNLOAD_DIR.glob("*.crdownload")):
            path = max(new_files, key=lambda p: p.stat().st_mtime)
            log.info("Downloaded: %s (%d bytes)", path, path.stat().st_size)
            return path
        time.sleep(2)
    raise TimeoutError("Download did not complete in time")


def extract_orders(path):
    """Step 6: read the CSV and return the order records."""
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        log.warning("CSV is empty: %s", path)
        return []

    header, data = rows[0], rows[1:]
    records = [
        # TikTok pads IDs with a trailing tab (anti-Excel-mangling) — strip it
        {k.strip(): v.strip() for k, v in zip(header, row) if v.strip()}
        for row in data
    ]
    order_ids = list(dict.fromkeys(r["Order ID"] for r in records if r.get("Order ID")))
    log.info("Export OK: %d row(s) -> %d unique order(s) [%s]",
             len(records), len(order_ids), path.name)
    return records


def match_orders_in_shopify(records):
    """Phase 2: find each TikTok order in Shopify via its TikTokOrderID tag."""
    from services.shopify import client as shopify

    order_ids = list(dict.fromkeys(r["Order ID"] for r in records if r.get("Order ID")))
    log_phase("PHASE 2/3: Match orders in Shopify")
    log.info("Looking up %d TikTok order ID(s) in Shopify by TikTokOrderID tag ...",
             len(order_ids))

    found, missing = [], []
    for oid in order_ids:
        try:
            order = shopify.find_order_by_tiktok_id(oid)
        except Exception as exc:
            log.error("  %s -> Shopify lookup failed: %s", oid, exc)
            missing.append(oid)
            continue
        if order:
            found.append({"tiktok_id": oid, "shopify_order": order})
            log.info("  %s -> %s (%s / %s)", oid, order["name"],
                     order["displayFinancialStatus"], order["displayFulfillmentStatus"])
        else:
            missing.append(oid)
            log.info("  %s -> NOT FOUND in Shopify", oid)

    log.info("Result: %d matched in Shopify, %d not found (of %d).",
             len(found), len(missing), len(order_ids))
    if missing:
        log.info("Not in Shopify (likely not synced yet): %s", ", ".join(missing))
    return found, missing


# ===========================================================================
# Phase 3 — my-fulfilment.com lookup (Nic. Oud portal)
# ===========================================================================
#
# This runs over the portal's HTTP API via services/fulfilment/client.py:
# one form-POST login, then plain GET requests. No browser, no page loads,
# no waiting on the Filters panel to render.
#
# Everything below the "DEPRECATED" banner further down is the original
# Selenium implementation. It is kept intact and still works — set
# TIKTOK_FULFILMENT_USE_BROWSER=1 to fall back to it — but it is no longer
# the path this workflow takes.
#
# The two produce the same dict shape on purpose; see fulfilment_lookup_api.


def _use_browser_fulfilment():
    """True if the deprecated Selenium fulfilment path is explicitly requested."""
    return os.environ.get("TIKTOK_FULFILMENT_USE_BROWSER", "").strip().lower() in (
        "1", "true", "yes", "on")


_myf_client = None


def get_fulfilment_client():
    """
    Return a logged-in MyFulfilmentClient, creating it on first use.

    One session is reused for every lookup in the run — the portal serialises
    requests sharing a PHPSESSID anyway, so a second session would buy nothing
    here (see services/fulfilment/pool.py for why parallelism needs separate
    logins). The cookie is cached to fulfilment_session.cookies, so repeat runs
    usually skip the login round-trip entirely.
    """
    global _myf_client
    if _myf_client is None:
        from services.fulfilment import client as myf
        c = myf.MyFulfilmentClient()
        if not c.is_logged_in():
            log.info("Logging in to my-fulfilment.com ...")
            c.login()
            log.info("my-fulfilment.com login successful.")
        else:
            log.info("Reusing cached my-fulfilment.com session.")
        _myf_client = c
    return _myf_client


def close_fulfilment_client():
    """
    Release the shared client at the end of a run.

    Mirrors MyFulfilmentClient.__exit__: the session cookie MUST be saved
    before closing, otherwise the next run pays for a fresh login. There is
    no close() method on the client — don't "simplify" this to one.
    """
    global _myf_client
    if _myf_client is not None:
        try:
            _myf_client.save_session()
            _myf_client.session.close()
        except Exception as exc:
            log.debug("Ignoring error while closing fulfilment session: %s", exc)
        _myf_client = None


def fulfilment_lookup_api(reference):
    """
    Look up an order in my-fulfilment.com over HTTP.

    Returns exactly the same dict that the deprecated Selenium
    fulfilment_lookup() returned, so log_fulfilment_result() and
    run_upload_phase() consume it unchanged:

        reference        the reference we searched for
        order_row        grid row, keyed by the portal's column headings
        detail_url       link to the order's /show or /edit page
        order_lines      list of dicts, keyed by the detail table headings
        packages         list of dicts, keyed by the detail table headings
        tracking         [{"text", "url"}] carrier links only
        carrier          Shipper of the first tracked package
        tracking_number  its TNT code
        shipped          whether a tracking number was found

    Returns None when no order matches, same as before.
    """
    client = get_fulfilment_client()
    row = client.find_order(reference)
    if row is None:
        log.warning("No fulfilment order found for reference %s", reference.lstrip("#"))
        return None

    # Keys mirror the portal's grid headings, because log_fulfilment_result()
    # reads them by those names.
    summary = {
        "reference": reference,
        "order_row": {
            "Nic. Oud reference": row.wms_reference,
            "Reference": row.reference,
            "Status": row.status,
            "Name": row.name,
            "Address": row.address,
            "City": row.city,
            "Country": row.country,
            "Lines": row.lines,
            "Created at": row.created_at,
            "Modified at": row.modified_at,
        },
    }

    if not row.detail_url:
        log.warning("Fulfilment order row found for %s but no detail link",
                    reference.lstrip("#"))
        return summary

    detail = client.get_order_detail(row)
    summary["detail_url"] = detail.url
    summary["order_lines"] = [
        {
            "Article code": line.article_code,
            "Article description": line.description,
            "Value": line.value,
            "Serial numbers": line.serial_numbers,
            "Ordered": line.ordered,
            "Shipped": line.shipped,
            "Open": line.open,
        }
        for line in detail.lines
    ]
    summary["packages"] = [
        {
            "Shipping date": pkg.shipping_date,
            "Shipper": pkg.shipper,
            "Shipping method": pkg.shipping_method,
            "Weight": pkg.weight,
            "TNT-Code": pkg.tnt_code,
        }
        for pkg in detail.packages
    ]
    # The portal renders each T&T code as a real carrier link carrying the
    # postcode/country params the carrier needs — keep those, drop any link
    # that points back into the portal (those are column sort links).
    summary["tracking"] = [
        {"text": pkg.tnt_code, "url": pkg.tnt_url}
        for pkg in detail.packages
        if pkg.tnt_url and "my-fulfilment.com" not in pkg.tnt_url
    ]

    # First tracked package wins — these two feed the TikTok upload in phase 4.
    # NOTE: is_tracked also rejects a literal "-", which the portal writes for
    # non-tracked delivery modes. The old Selenium path treated "-" as a real
    # tracking number and would have uploaded it to TikTok.
    summary["carrier"] = ""
    summary["tracking_number"] = ""
    for pkg in detail.packages:
        if pkg.is_tracked:
            summary["carrier"] = (pkg.shipper or "").strip()
            summary["tracking_number"] = pkg.tnt_code.strip()
            break
    summary["shipped"] = bool(summary["tracking_number"])
    return summary


# ===========================================================================
# DEPRECATED — Selenium my-fulfilment.com navigation
# ===========================================================================
#
# Superseded by fulfilment_lookup_api() above, which talks to the portal over
# HTTP instead of driving it through Chrome. Kept because it still works and
# is the documented fallback (TIKTOK_FULFILMENT_USE_BROWSER=1) if the portal
# changes in a way the HTTP client can't follow.
#
# Everything from here to run_fulfilment_phase() is part of that old path:
#   _load_fulfilment_creds, _fulfilment_session_ok, _fulfilment_auto_login,
#   ensure_fulfilment_logged_in, _read_table, _load_fulfilment_page,
#   fulfilment_lookup
#
# Note it has its own credential loading and its own login form handling,
# separate from services/fulfilment/client.py.

FULFILMENT_LOGIN_URL = "https://www.my-fulfilment.com/login"


def _load_fulfilment_creds():
    """DEPRECATED (browser path). Read my-fulfilment.com login from env vars or fulfilment.env."""
    email = os.getenv("MYFULFILMENT_EMAIL")
    password = os.getenv("MYFULFILMENT_PASSWORD")
    env_file = PROJECT_ROOT / "fulfilment.env"
    if (not email or not password) and env_file.exists():
        for line in env_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                if k.strip() == "MYFULFILMENT_EMAIL" and not email:
                    email = v.strip()
                elif k.strip() == "MYFULFILMENT_PASSWORD" and not password:
                    password = v.strip()
    return email, password


def _fulfilment_session_ok(driver):
    """DEPRECATED (browser path). True if the orders page is loaded with a valid session."""
    try:
        body = driver.find_element(By.TAG_NAME, "body").text
    except WebDriverException:
        body = ""
    if "HTTP ERROR 5" in body or "isn't working" in body:
        return False  # stale session is served as a 500 by the portal
    if "/login" in driver.current_url.lower():
        return False
    if find_visible(driver, "//input[@type='password']") is not None:
        return False
    return True


# DEPRECATED (browser path) — see services/fulfilment/client.py login().
def _fulfilment_auto_login(driver):
    """Fill the login form from creds, tick 'Remember me', submit."""
    from selenium.webdriver.common.keys import Keys

    email, password = _load_fulfilment_creds()
    if not email or not password:
        raise RuntimeError(
            "my-fulfilment.com session expired and no credentials available. "
            "Set MYFULFILMENT_EMAIL / MYFULFILMENT_PASSWORD in fulfilment.env."
        )

    _load_fulfilment_page(driver, FULFILMENT_LOGIN_URL)
    time.sleep(3)

    email_in = find_visible(driver, "//input[@type='email']") or \
        find_visible(driver, "//input[@type='text']")
    pass_in = find_visible(driver, "//input[@type='password']")
    if not email_in or not pass_in:
        screenshot(driver, "fulfilment_login_no_fields")
        raise RuntimeError("Could not locate my-fulfilment.com login fields")

    # JS-clear first: the browser autofills these, and typing on top would
    # duplicate the value (e.g. 'a@b.coma@b.com').
    for el, val in ((email_in, email), (pass_in, password)):
        driver.execute_script("arguments[0].value='';", el)
        el.click()
        el.send_keys(Keys.CONTROL, "a")
        el.send_keys(Keys.DELETE)
        el.send_keys(val)

    # Tick "Remember me" so the session persists and we avoid re-logging in.
    remember = find_visible(driver, "//input[@type='checkbox']")
    if remember and not remember.is_selected():
        try:
            js_click(driver, remember)
        except WebDriverException:
            pass
    if remember and remember.is_selected():
        log.info("'Remember me' checked.")

    submit = find_visible(driver, "//button[@type='submit']") or \
        find_visible(driver, "//button[contains(translate(., 'LOGIN', 'login'), 'login')]") or \
        find_visible(driver, "//input[@type='submit']")
    if submit:
        js_click(driver, submit)
    else:
        pass_in.send_keys(Keys.RETURN)
    log.info("Submitted my-fulfilment.com login for %s", email)
    time.sleep(6)


def ensure_fulfilment_logged_in(driver, headless):
    """DEPRECATED (browser path) — the HTTP client logs itself in.

    Ensure a valid my-fulfilment.com session, auto-logging in if needed.

    The portal serves a stale session as an HTTP 500 on /orders/ rather than
    a clean login redirect, so we detect that and re-authenticate from creds.
    """
    _load_fulfilment_page(driver, FULFILMENT_ORDERS_URL)
    time.sleep(5)

    if _fulfilment_session_ok(driver):
        log.info("Already logged in to my-fulfilment.com.")
        return

    log.info("my-fulfilment.com session invalid/expired — logging in automatically.")
    _fulfilment_auto_login(driver)

    _load_fulfilment_page(driver, FULFILMENT_ORDERS_URL)
    time.sleep(4)
    if _fulfilment_session_ok(driver):
        log.info("my-fulfilment.com login successful, session saved to chrome_profile.")
        return
    screenshot(driver, "fulfilment_login_failed")
    raise RuntimeError("my-fulfilment.com login failed (check credentials or portal status)")


def _read_table(driver, heading):
    """DEPRECATED (browser path). Parse the first table following a section heading into row dicts."""
    table = find_visible(driver, f"//*[normalize-space(text())='{heading}']/following::table[1]")
    if not table:
        return [], []
    headers = [th.text.strip() for th in table.find_elements(By.XPATH, ".//th")]
    rows = []
    for tr in table.find_elements(By.XPATH, ".//tbody/tr"):
        cells = tr.find_elements(By.XPATH, "./td")
        if cells:
            rows.append({h: c.text.strip() for h, c in zip(headers, cells)})
    links = [{"text": a.text.strip(), "url": a.get_attribute("href")}
             for a in table.find_elements(By.XPATH, ".//a")
             if a.text.strip()]
    return rows, links


def _load_fulfilment_page(driver, url, attempts=4):
    """DEPRECATED (browser path). Navigate to a my-fulfilment.com page, retrying on a transient HTTP 500.

    The portal intermittently returns a "This page isn't working / HTTP ERROR
    500" page; a reload after a short wait usually fixes it.
    """
    for attempt in range(1, attempts + 1):
        driver.get(url)
        time.sleep(3)
        body = ""
        try:
            body = driver.find_element(By.TAG_NAME, "body").text
        except WebDriverException:
            pass
        if "isn't working" not in body and "HTTP ERROR 5" not in body \
                and "unable to handle this request" not in body:
            return True
        log.warning("my-fulfilment.com returned a server error (attempt %d/%d) — "
                    "retrying in %ds", attempt, attempts, 5 * attempt)
        time.sleep(5 * attempt)
    log.error("my-fulfilment.com still erroring after %d attempts: %s", attempts, url)
    return False


def fulfilment_lookup(driver, reference):
    """DEPRECATED — superseded by fulfilment_lookup_api(), which does the same
    lookup over HTTP and returns the identical dict shape.

    Filter my-fulfilment.com orders by reference and pull full details.

    Returns a dict with the result row, order lines, packages and tracking
    link(s), or None if no order matches the reference.
    """
    ref = reference.lstrip("#")
    if not _load_fulfilment_page(driver, FULFILMENT_ORDERS_URL):
        raise RuntimeError("my-fulfilment.com orders page unavailable (HTTP 500)")

    # The Reference input lives in the collapsible Filters panel.
    ref_input = find_visible(driver, "//input[@placeholder='Reference']")
    if not ref_input:
        toggle = find_visible(driver, "//*[normalize-space(text())='Filters']")
        if toggle:
            js_click(driver, toggle)
        ref_input = wait_for(driver, "//input[@placeholder='Reference']",
                             20, "Reference filter input")
    ref_input.clear()
    ref_input.send_keys(ref)
    apply_btn = wait_for(
        driver,
        "//button[normalize-space()='Apply filters'] | //input[@value='Apply filters']",
        10, "Apply filters button",
    )
    js_click(driver, apply_btn)
    time.sleep(4)

    row = find_visible(driver, f"//tr[td[contains(normalize-space(), '{ref}')]]")
    if not row:
        log.warning("No fulfilment order found for reference %s", ref)
        return None

    header_cells = [th.text.strip() for th in driver.find_elements(
        By.XPATH, f"//tr[td[contains(normalize-space(), '{ref}')]]/ancestor::table[1]//th")]
    cells = [td.text.strip() for td in row.find_elements(By.XPATH, "./td")]
    summary = {
        "reference": reference,
        "order_row": {h: c for h, c in zip(header_cells, cells) if h},
    }

    # The grid row carries its detail page in data-href (no <a> elements).
    detail_url = row.get_attribute("data-href")
    if not detail_url:
        for a in row.find_elements(By.XPATH, ".//a[contains(@href, '/orders/')]"):
            detail_url = a.get_attribute("href")
            break
    if not detail_url:
        log.warning("Fulfilment order row found for %s but no detail link", ref)
        return summary

    detail_url = detail_url.replace("/edit", "/show")
    if detail_url.startswith("/"):
        detail_url = "https://www.my-fulfilment.com" + detail_url
    _load_fulfilment_page(driver, detail_url)
    summary["detail_url"] = detail_url
    summary["order_lines"], _ = _read_table(driver, "Order lines")
    summary["packages"], package_links = _read_table(driver, "Packages")
    # Keep only carrier links (the table headers are sort links back into
    # the portal — only external URLs are actual tracking links).
    summary["tracking"] = [
        l for l in package_links
        if l["url"] and "my-fulfilment.com" not in l["url"]
    ]

    # Extract the shipped package's carrier + tracking number (the inputs
    # phase 4 needs). A "Confirmed"/unshipped order has no package yet.
    summary["carrier"] = ""
    summary["tracking_number"] = ""
    for pkg in summary["packages"]:
        tnt = (pkg.get("TNT-Code") or "").strip()
        if tnt:
            summary["carrier"] = (pkg.get("Shipper") or "").strip()
            summary["tracking_number"] = tnt
            break
    summary["shipped"] = bool(summary["tracking_number"])
    return summary


def log_fulfilment_result(tiktok_id, shopify_name, info):
    """Log one consolidated end-summary block for an order."""
    log.info("-" * 60)
    log.info("TikTok order : %s", tiktok_id)
    log.info("Shopify order: %s", shopify_name)
    if not info:
        log.info("Fulfilment   : NOT FOUND for reference %s", shopify_name)
        return
    row = info.get("order_row", {})
    log.info("Fulfilment   : Nic.Oud ref %s | status %s | %s, %s, %s",
             row.get("Nic. Oud reference", "?"), row.get("Status", "?"),
             row.get("Name", "?"), row.get("City", "?"), row.get("Country", "?"))
    for line in info.get("order_lines", []):
        log.info("Order line   : %s x%s (%s) shipped=%s open=%s",
                 line.get("Article code", "?"), line.get("Ordered", "?"),
                 line.get("Article description", ""), line.get("Shipped", "?"),
                 line.get("Open", "?"))
    if info.get("shipped"):
        for pkg in info.get("packages", []):
            if (pkg.get("TNT-Code") or "").strip():
                log.info("Package      : shipped %s via %s (%s), weight %s, TNT-code %s",
                         pkg.get("Shipping date", "?"), pkg.get("Shipper", "?"),
                         pkg.get("Shipping method", "?"), pkg.get("Weight", "?"),
                         pkg.get("TNT-Code", "?"))
        for t in info.get("tracking", []):
            log.info("Tracking     : %s -> %s", t["text"], t["url"])
    else:
        log.info("Package      : NOT SHIPPED YET — no tracking number "
                 "(status %s); will be skipped for TikTok upload.",
                 info.get("order_row", {}).get("Status", "?"))
    if info.get("detail_url"):
        log.info("Detail page  : %s", info["detail_url"])


def run_fulfilment_phase(driver, headless, found):
    """Phase 3: look up each Shopify-matched order in my-fulfilment.com.

    Uses the portal's HTTP API. `driver` and `headless` are still accepted so
    callers don't change, and are only used by the deprecated browser path
    (TIKTOK_FULFILMENT_USE_BROWSER=1).
    """
    log_phase("PHASE 3/3: Look up tracking in my-fulfilment.com")
    log.info("Checking %d matched order(s) for tracking ...", len(found))

    use_browser = _use_browser_fulfilment()
    if use_browser:
        log.warning("TIKTOK_FULFILMENT_USE_BROWSER=1 — using the DEPRECATED "
                    "Selenium fulfilment path.")
        ensure_fulfilment_logged_in(driver, headless)
    else:
        log.info("Fulfilment lookups via my-fulfilment.com API (no browser).")

    results = []
    try:
        for item in found:
            name = item["shopify_order"]["name"]
            try:
                if use_browser:
                    info = fulfilment_lookup(driver, name)
                else:
                    info = fulfilment_lookup_api(name)
            except Exception as exc:
                log.error("Fulfilment lookup failed for %s: %s", name, exc)
                if use_browser:
                    screenshot(driver, f"fulfilment_fail_{name.lstrip('#')}")
                info = None
            results.append((item["tiktok_id"], name, info))
    finally:
        if not use_browser:
            close_fulfilment_client()

    shipped = sum(1 for _, _, info in results if info and info.get("shipped"))
    log_phase("FINAL SUMMARY")
    log.info("%d order(s) processed: %d shipped (tracking found), %d not yet shipped/found.",
             len(results), shipped, len(results) - shipped)
    for tiktok_id, name, info in results:
        log_fulfilment_result(tiktok_id, name, info)
    log.info("=" * 70)
    return results


def run_upload_phase(driver, results, dry_run=None):
    """Phase 4: submit found tracking back to TikTok Seller Center.

    Builds the shipment list from Phase 3 results (only orders that are
    actually shipped and have a tracking number), then submits each one
    through the Seller Center UI:

        To ship list -> Add tracking info -> [Combine orders] -> Submit parcel

    Set TIKTOK_UPLOAD_USE_TEMPLATE=1 to use the DEPRECATED template
    download/fill/upload path instead.

    dry_run: fill everything but stop before the final Submit click. Defaults
    to the TIKTOK_UPLOAD_DRY_RUN env var, so a scheduled run submits for real
    while a manual verification run can stop short.
    """
    log_phase("PHASE 4/4: Submit tracking to TikTok Seller Center")
    shipments = [
        {
            "order_id": str(tiktok_id),
            "provider": tiktok_provider(info.get("carrier")),
            "tracking": info["tracking_number"],
        }
        for tiktok_id, _name, info in results
        if info and info.get("shipped") and info.get("tracking_number")
    ]
    if not shipments:
        log.info("No shipped orders with tracking — nothing to upload to TikTok.")
        return None

    log.info("Submitting tracking for %d order(s) to TikTok ...", len(shipments))
    for s in shipments:
        log.info("  order %s -> %s (%s)", s["order_id"], s["tracking"], s["provider"])

    if os.environ.get("TIKTOK_UPLOAD_USE_TEMPLATE", "").strip().lower() in (
            "1", "true", "yes", "on"):
        log.warning("TIKTOK_UPLOAD_USE_TEMPLATE=1 — using the DEPRECATED "
                    "template upload path.")
        template = download_ship_template(driver)
        filled = fill_ship_template(template, shipments)
        ok, verdict = upload_ship_file(driver, filled)
        log.info("TikTok upload %s: %s", "SUCCEEDED" if ok else "did not confirm", verdict)
        return ok

    if dry_run is None:
        dry_run = os.environ.get("TIKTOK_UPLOAD_DRY_RUN", "").strip().lower() in (
            "1", "true", "yes", "on")
    return submit_tracking_via_ui(driver, shipments, dry_run=dry_run)


# ---------------------------------------------------------------------------
# Phase 4 — per-order "Add tracking info" flow (Seller Center UI)
# ---------------------------------------------------------------------------

ORDER_ID_RE = r"\b\d{15,}\b"

# Collect order ids from a subtree (or the whole document when scope is null).
#
# Do NOT regex textContent for this. Concatenated textContent glues an order id
# to its neighbours — "...Shipped by seller576933471526951296" then "1 item" —
# so \b\d{15,}\b finds nothing (no word boundary between "r" and "5") or worse,
# matches a corrupted "5769334715269512961" made of the id plus the next digit.
# Seller Center puts every id in its own leaf element (class order_id_number),
# so match leaves whose entire text is the id and nothing else.
_ORDER_IDS_JS = """
var scope = arguments[0] || document;
var out = [];
var nodes = scope.querySelectorAll('*');
for (var i = 0; i < nodes.length; i++) {
    var n = nodes[i];
    if (n.children.length) continue;                 // leaves only
    var t = (n.textContent || '').trim();
    if (/^\\d{15,}$/.test(t)) out.push(t);
}
return Array.from(new Set(out));
"""


def order_ids_in(driver, scope=None):
    """Return the distinct order ids rendered inside `scope` (default: page)."""
    return driver.execute_script(_ORDER_IDS_JS, scope) or []

# The order-id element is a ~21px text line while list rows sit ~104px apart,
# and the action button renders a little below that line. Half the row pitch
# is therefore the widest a match can be while still being unambiguous.
ROW_MATCH_TOLERANCE_PX = 50


def _row_container(driver, el):
    """
    Climb from `el` to the element that represents its whole table row.

    Seller Center renders these grids as nested divs as often as real <tr>s, so
    anchoring on a tag name is unreliable.

    Do NOT stop at the first ancestor containing an order id — that is just the
    Order cell, and the row's action buttons live in a different cell entirely.
    Instead keep climbing while the subtree still contains only THIS order id,
    and take the last one that holds. The first ancestor to pull in a second
    order id is the table body, so the one before it is the row.
    """
    return driver.execute_script(
        """
        var el = arguments[0];

        function ids(node) {
            var out = [];
            var nodes = node.querySelectorAll('*');
            for (var i = 0; i < nodes.length; i++) {
                var n = nodes[i];
                if (n.children.length) continue;
                var t = (n.textContent || '').trim();
                if (/^\\d{15,}$/.test(t)) out.push(t);
            }
            var self = (node.textContent || '').trim();
            if (!node.children.length && /^\\d{15,}$/.test(self)) out.push(self);
            return Array.from(new Set(out));
        }

        // The id this element belongs to (search up until we see exactly one).
        var own = null, probe = el;
        for (var i = 0; i < 6 && probe; i++) {
            var f = ids(probe);
            if (f.length === 1) { own = f[0]; break; }
            probe = probe.parentElement;
        }
        if (!own) return null;

        var node = probe, best = probe;
        for (var j = 0; j < 12 && node && node.parentElement; j++) {
            node = node.parentElement;
            var f2 = ids(node);
            if (f2.length === 1 && f2[0] === own) best = node;
            else if (f2.length > 1) break;   // reached the table body
        }
        return best;
        """,
        el,
    )


_FIND_BY_TEXT_JS = """
var scope = arguments[0] || document;
var pattern = arguments[1];
var isRegex = arguments[2];
var re = isRegex ? new RegExp(pattern) : null;

function norm(node) {
    return (node.textContent || '').replace(/\\s+/g, ' ').trim();
}
function hit(text) {
    return isRegex ? re.test(text) : text === pattern;
}

var out = [];
var nodes = scope.querySelectorAll('*');
for (var i = 0; i < nodes.length; i++) {
    var node = nodes[i];
    if (node.offsetParent === null) continue;              // not rendered
    var text = norm(node);
    if (!text || text.length > 120) continue;              // prune containers
    if (!hit(text)) continue;
    // Keep only the innermost element carrying the text: TikTok wraps button
    // labels in a span, and clicking the span bubbles to the button anyway.
    var inner = false;
    var kids = node.querySelectorAll('*');
    for (var j = 0; j < kids.length; j++) {
        if (hit(norm(kids[j]))) { inner = true; break; }
    }
    if (!inner) out.push(node);
}
return out.length ? out[0] : null;
"""


def find_by_text(driver, pattern, scope=None, regex=False):
    """
    Return the innermost visible element whose text matches `pattern`.

    Prefer this over an XPath like
        //button[contains(., 'X')][not(.//*[contains(., 'X')])]
    which cannot match TikTok's buttons at all: the <button> fails the
    not-a-descendant test because its label sits in an inner <span>, and the
    <span> fails the tag test. Clicking the inner element bubbles to the
    button, so matching the innermost node is both simpler and correct.
    """
    return driver.execute_script(_FIND_BY_TEXT_JS, scope, pattern, regex)


def _row_action_button(driver, row, label):
    """
    Find a row's action button by VERTICAL POSITION, not DOM ancestry.

    The action column is sticky, so Seller Center renders it as a separate
    parallel table pinned to the right edge. The button is therefore never a
    descendant of the row element on the left — searching inside the row finds
    nothing no matter how far up you climb.

    What is reliable is geometry: the button sits on the same horizontal band
    as its row. So take the row's vertical extent and pick the button whose
    centre falls inside it. This also disambiguates when several rows have
    rendered their buttons, which happens because they do not disappear once
    hovered.
    """
    return driver.execute_script(
        """
        var row = arguments[0], label = arguments[1], tolerance = arguments[2];
        var r = row.getBoundingClientRect();
        var rowMid = r.top + r.height / 2;

        var best = null, bestDist = Infinity;
        document.querySelectorAll('button, a, div[role="button"], span').forEach(
            function(node) {
                if ((node.textContent || '').trim() !== label) return;
                if (node.offsetParent === null) return;             // not visible
                if (node.querySelector('button, a, div[role="button"]')) return;  // wrapper
                var b = node.getBoundingClientRect();
                if (b.width === 0 || b.height === 0) return;
                var dist = Math.abs((b.top + b.height / 2) - rowMid);
                if (dist < bestDist) { bestDist = dist; best = node; }
            });
        return bestDist <= tolerance ? best : null;
        """,
        row, label, ROW_MATCH_TOLERANCE_PX,
    )


def _find_order_row(driver, order_id):
    """Return the To-ship list row element for `order_id`, or None."""
    for el in driver.find_elements(
            By.XPATH, f"//*[normalize-space(text())='{order_id}']"):
        try:
            row = _row_container(driver, el)
            if row is not None:
                return row
        except WebDriverException:
            continue
    return None


def _handle_combine_dialog(driver, order_id, known_orders):
    """
    Deal with the optional "Combine orders" dialog.

    TikTok shows it when the same buyer/address has other shippable orders.
    With nothing to combine it never appears, so it must be treated as
    optional — never wait on it as if it were guaranteed.

    Which button we press depends on whether we can cover everything in the
    parcel. The orders grouped here are usually NOT all shipped yet: the buyer
    ordered three times, Nic. Oud has shipped one. Combining them anyway would
    put this order's tracking number on the other two, which are real customer
    orders. So we only combine when every order in the dialog has its own
    tracking; otherwise we take "Continue without combining" and pick the
    others up on a later run, once they ship.

    Set TIKTOK_ALWAYS_COMBINE=1 to always combine regardless.
    """
    # Detect the dialog by its buttons rather than its title: "Combine orders"
    # as a title is easy to confuse with other text, whereas these two labels
    # exist only here.
    anchor = find_by_text(driver, "Continue without combining")
    if anchor is None:
        log.info("No Combine orders dialog (nothing to combine) — continuing.")
        return

    # Scope to the dialog before reading order ids: the page behind it is full
    # of 15+ digit product ids, and scraping document.body picks all of them up,
    # making every order look "unknown".
    #
    # Climb from the BUTTON until the ancestor actually contains order ids.
    # Climbing from the title instead lands on a container that holds the
    # buttons but not the parcel rows, which silently yields zero ids — and
    # zero unknown ids reads as "safe to combine", the exact wrong answer.
    dialog = driver.execute_script(
        "return arguments[0].closest('.p-modal, [role=\"dialog\"]') "
        "|| arguments[0].closest('.p-modal-wrapper');",
        anchor,
    )
    if dialog is None:
        screenshot(driver, f"combine_dialog_unparsed_{order_id}")
        raise RuntimeError("Combine dialog found but its container could not be resolved")

    ids = []
    deadline = time.time() + 20
    while time.time() < deadline:
        ids = order_ids_in(driver, dialog)
        if ids:
            break
        time.sleep(1)

    listed = list(dict.fromkeys(ids))
    if not listed:
        screenshot(driver, f"combine_dialog_no_orders_{order_id}")
        raise RuntimeError("Combine dialog listed no order ids — refusing to "
                           "decide whether combining is safe")

    unknown = [i for i in listed if i not in known_orders]

    always = os.environ.get("TIKTOK_ALWAYS_COMBINE", "").strip().lower() in (
        "1", "true", "yes", "on")

    if unknown and not always:
        log.warning("Combine orders dialog lists %d order(s) with no tracking of "
                    "their own (%s) — combining would put %s's tracking on them. "
                    "Choosing 'Continue without combining'.",
                    len(unknown), ", ".join(unknown[:5]), order_id)
        button_text = "Continue without combining"
    else:
        if unknown and always:
            log.warning("TIKTOK_ALWAYS_COMBINE=1 — combining despite %d order(s) "
                        "without their own tracking.", len(unknown))
        else:
            log.info("Combine orders dialog: all %d order(s) have tracking — "
                     "combining.", len(listed))
        button_text = "Combine orders and continue"

    btn = find_by_text(driver, button_text, scope=dialog)
    if btn is None:
        screenshot(driver, f"combine_dialog_{order_id}")
        raise RuntimeError(f"Combine dialog shown but '{button_text}' button not found")

    log.info("Clicking '%s'.", button_text)
    js_click(driver, btn)
    time.sleep(5)


def open_add_tracking(driver, order_id, known_orders=()):
    """
    Click "Add tracking info" for one order and land on the tracking page.

    known_orders: order ids we hold tracking for — used to decide whether
    combining is safe (see _handle_combine_dialog).

    Returns True if we reached the Add tracking info page.
    """
    driver.get(ORDERS_URL)
    time.sleep(6)

    row = _find_order_row(driver, order_id)
    if row is None:
        log.error("Order %s is not on the To ship list — cannot add tracking.", order_id)
        screenshot(driver, f"order_not_listed_{order_id}")
        return False

    # The button only renders on row hover, so fire the hover first. It also
    # sits next to a dropdown caret; take the button itself, not the caret.
    driver.execute_script(
        """
        var row = arguments[0];
        row.scrollIntoView({block: 'center'});
        ['mouseover', 'mouseenter', 'mousemove'].forEach(function(t) {
            row.dispatchEvent(new MouseEvent(t, {bubbles: true}));
        });
        """,
        row,
    )
    time.sleep(1)

    btn = _row_action_button(driver, row, "Add tracking info")
    if btn is None:
        log.error("No 'Add tracking info' button in the row for %s "
                  "(order may not be in an addable state).", order_id)
        screenshot(driver, f"no_add_tracking_btn_{order_id}")
        return False

    js_click(driver, btn)
    log.info("Clicked 'Add tracking info' for order %s", order_id)
    time.sleep(4)

    _handle_combine_dialog(driver, order_id, known_orders)

    if not find_visible(driver, "//*[contains(normalize-space(text()), 'Add tracking info')]"):
        log.error("Did not reach the Add tracking info page for %s", order_id)
        screenshot(driver, f"no_tracking_page_{order_id}")
        return False

    # Wait for the parcel table to render before reading it. The heading and
    # the page shell appear first; reading now returns only internal ids from
    # the loading state and the order looks absent when it is simply not drawn
    # yet.
    deadline = time.time() + 30
    while time.time() < deadline:
        if driver.find_elements(By.XPATH, "//input[@placeholder='Enter tracking ID']"):
            break
        time.sleep(1)
    else:
        log.error("Tracking table never rendered for %s", order_id)
        screenshot(driver, f"no_tracking_table_{order_id}")
        return False

    # Confirm we opened OUR order. Cheap, and the thing that makes the
    # geometry-based button match safe to rely on.
    page_ids = order_ids_in(driver)
    if order_id not in page_ids:
        log.error("Add tracking info page does not list order %s (found %s) — "
                  "wrong row was opened, aborting.", order_id, sorted(set(page_ids))[:5])
        screenshot(driver, f"wrong_order_page_{order_id}")
        return False
    return True


def _select_shipping_provider(driver, row, provider):
    """
    Pick `provider` in a row's Shipping provider dropdown.

    It is a custom widget, not a <select>, so the options render in a portal
    appended to <body> — they can't be found by searching inside the row.
    """
    control = driver.execute_script(
        """
        var row = arguments[0];
        var nodes = row.querySelectorAll('input, div, span');
        for (var i = 0; i < nodes.length; i++) {
            var n = nodes[i];
            var txt = (n.getAttribute('placeholder') || n.textContent || '').trim();
            if (txt.indexOf('Select shipping provider') !== -1) return n;
        }
        return null;
        """,
        row,
    )
    if control is None:
        raise RuntimeError("Shipping provider control not found in row")

    js_click(driver, control)
    time.sleep(2)

    # Some builds render a searchable combobox — typing narrows a long list.
    try:
        if control.tag_name == "input":
            control.send_keys(provider)
            time.sleep(2)
    except WebDriverException:
        pass

    log.info("    picking provider option '%s' ...", provider)
    deadline = time.time() + 20
    option = None
    while time.time() < deadline:
        option = find_by_text(driver, provider)
        if option is not None:
            break
        time.sleep(1)
    if option is None:
        screenshot(driver, "no_provider_option")
        raise TimeoutError(f"Shipping provider option '{provider}' not found")
    js_click(driver, option)
    time.sleep(1)
    log.info("    provider set to %s", provider)


def _tracking_rows(driver):
    """
    Return [(row_element, tracking_input, [order_ids_in_row])] for the page.

    A combined parcel puts several orders on the page, so each row is matched
    to its own order rather than assuming a single one.
    """
    rows = []
    for inp in driver.find_elements(
            By.XPATH, "//input[@placeholder='Enter tracking ID']"):
        try:
            if not inp.is_displayed():
                continue
            row = _row_container(driver, inp)
            if row is None:
                continue
            ids = order_ids_in(driver, row)
            rows.append((row, inp, ids))
        except WebDriverException:
            continue
    return rows


class UnknownParcelOrder(RuntimeError):
    """A parcel row belongs to an order we have no tracking for."""


def fill_tracking_page(driver, by_order, fallback):
    """
    Fill every row on the Add tracking info page.

    by_order: {order_id: shipment} for everything we have tracking for.
    fallback: the shipment we opened the page with. Used only for a row whose
              order ids could not be read at all — never to invent a tracking
              number for a known-but-different order.

    Raises UnknownParcelOrder if a row belongs to an order we have no tracking
    for. Submitting one order's tracking against another customer's order is
    worse than not submitting at all, so the caller skips the whole parcel and
    retries on a later run.

    Returns the set of order ids covered.
    """
    rows = _tracking_rows(driver)
    if not rows:
        raise RuntimeError("No tracking ID inputs found on the Add tracking info page")

    log.info("  Add tracking info page has %d parcel row(s)", len(rows))

    # Check every row before typing anything, so we never half-fill a page.
    for _row, _inp, ids in rows:
        if ids and not any(i in by_order for i in ids):
            raise UnknownParcelOrder(
                f"parcel row {ids} has no tracking of its own — refusing to "
                f"submit {fallback['tracking']} against it")

    filled = set()
    for row, inp, ids in rows:
        shipment = next((by_order[i] for i in ids if i in by_order), None) or fallback
        inp.clear()
        inp.send_keys(shipment["tracking"])
        log.info("    order(s) %s -> tracking %s", ids or "?", shipment["tracking"])
        _select_shipping_provider(driver, row, shipment["provider"])
        # Receipt ID is optional — deliberately left blank.
        filled.update(i for i in ids if i in by_order)

    return filled


def submit_parcels(driver, dry_run):
    """Click "Submit N parcel(s)" and confirm. Returns (ok, verdict)."""
    # "Submit 1 parcel" / "Submit 3 parcels" — the count varies, so match a pattern.
    submit = find_by_text(driver, r"^Submit \d+ parcels?$", regex=True)
    if submit is None:
        screenshot(driver, "no_submit_button")
        return False, "Submit parcel button not found"

    label = (submit.text or "Submit").strip()
    if dry_run:
        screenshot(driver, "dry_run_before_submit")
        log.warning("DRY RUN — everything is filled in but '%s' was NOT clicked.", label)
        return True, f"dry run (stopped before '{label}')"

    js_click(driver, submit)
    log.info("  Clicked '%s'", label)
    time.sleep(5)

    # Some flows raise a confirmation dialog before committing.
    confirm = find_by_text(driver, r"^(Confirm|Submit|OK)$", regex=True)
    if confirm is not None:
        log.info("  Confirmation dialog shown — confirming.")
        js_click(driver, confirm)
        time.sleep(5)

    body = ""
    try:
        body = driver.find_element(By.TAG_NAME, "body").text.lower()
    except WebDriverException:
        pass
    if "success" in body or "submitted" in body:
        screenshot(driver, "submit_success")
        return True, "submitted"
    if "error" in body or "failed" in body or "invalid" in body:
        screenshot(driver, "submit_error")
        return False, "page reported an error"

    # Leaving the tracking page is itself the success signal in most builds.
    left_page = not find_visible(
        driver, "//input[@placeholder='Enter tracking ID']")
    screenshot(driver, "submit_result")
    if left_page:
        return True, "submitted (tracking page closed)"
    return False, "no confirmation after submit"


def submit_tracking_via_ui(driver, shipments, dry_run=False):
    """
    Phase 4: submit tracking one order at a time through the Seller Center UI.

    Combining can cover several orders in a single pass, so anything already
    submitted is skipped rather than opened again.
    """
    by_order = {s["order_id"]: s for s in shipments}
    done, failed = set(), []

    for shipment in shipments:
        order_id = shipment["order_id"]
        if order_id in done:
            log.info("Order %s already submitted as part of a combined parcel — skipping.",
                     order_id)
            continue

        log.info("-" * 60)
        log.info("Order %s: tracking %s via %s",
                 order_id, shipment["tracking"], shipment["provider"])
        try:
            if not open_add_tracking(driver, order_id, known_orders=set(by_order)):
                failed.append(order_id)
                continue
            try:
                covered = fill_tracking_page(driver, by_order, shipment)
            except UnknownParcelOrder as exc:
                log.warning("Order %s: skipped without submitting — %s. "
                            "It will be retried once the other order(s) ship.",
                            order_id, exc)
                screenshot(driver, f"unknown_parcel_{order_id}")
                failed.append(order_id)
                continue
            ok, verdict = submit_parcels(driver, dry_run)
            if ok:
                done.update(covered or {order_id})
                done.add(order_id)
                log.info("Order %s: %s (covered %d order(s))",
                         order_id, verdict, len(covered or {order_id}))
            else:
                failed.append(order_id)
                log.error("Order %s: submit failed — %s", order_id, verdict)
        except Exception as exc:
            log.exception("Order %s: tracking submission failed: %s", order_id, exc)
            screenshot(driver, f"tracking_fail_{order_id}")
            failed.append(order_id)

    log.info("=" * 70)
    log.info("TikTok tracking submission: %d/%d order(s) done%s",
             len(done & set(by_order)), len(shipments),
             " (DRY RUN — nothing was actually submitted)" if dry_run else "")
    if failed:
        log.warning("Failed order(s): %s", ", ".join(failed))
    return not failed


# ---------------------------------------------------------------------------
# DEPRECATED — bulk template download / fill / upload
# ---------------------------------------------------------------------------
#
# Superseded by submit_tracking_via_ui() above, which drives the per-order
# "Add tracking info" flow instead of moving an xlsx around. Kept working and
# reachable via TIKTOK_UPLOAD_USE_TEMPLATE=1.


def download_ship_template(driver):
    """DEPRECATED (template path). Download the shipping-upload template."""
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    driver.execute_cdp_cmd(
        "Browser.setDownloadBehavior",
        {"behavior": "allow", "downloadPath": str(DOWNLOAD_DIR)},
    )
    before = {p.name for p in DOWNLOAD_DIR.glob("*.xlsx")}

    driver.get(SHIP_UPLOAD_URL)
    btn = wait_for(driver,
                   "//*[self::button or self::a or self::span or self::div]"
                   "[contains(normalize-space(.), 'Download template')]"
                   "[not(.//*[contains(normalize-space(.), 'Download template')])]",
                   60, "Download template button")
    js_click(driver, btn)
    log.info("Clicked Download template — waiting for the xlsx file.")

    deadline = time.time() + DOWNLOAD_TIMEOUT
    while time.time() < deadline:
        new = [p for p in DOWNLOAD_DIR.glob("*.xlsx") if p.name not in before]
        if new and not list(DOWNLOAD_DIR.glob("*.crdownload")):
            path = max(new, key=lambda p: p.stat().st_mtime)
            log.info("Template downloaded: %s", path.name)
            return path
        time.sleep(2)
    raise TimeoutError("Shipping template download did not complete")


def fill_ship_template(template_path, shipments):
    """Phase 4b: fill the template's mandatory columns, preserving format.

    shipments: list of {"order_id", "provider", "tracking"}.
    Only the 'Shipping info' sheet gets data rows (from row 4); all other
    sheets (provider list, meta_info_sheet) are left untouched.
    """
    import openpyxl

    wb = openpyxl.load_workbook(template_path)
    ws = wb["Shipping info"]
    headers = {ws.cell(row=2, column=c).value: c for c in range(1, ws.max_column + 1)}
    col_order = headers["Order ID"]
    col_provider = headers["Shipping provider name"]
    col_tracking = headers["Tracking ID"]

    for i, s in enumerate(shipments):
        r = 4 + i
        ws.cell(row=r, column=col_order, value=str(s["order_id"]))
        ws.cell(row=r, column=col_provider, value=s["provider"])
        ws.cell(row=r, column=col_tracking, value=str(s["tracking"]))

    filled = template_path.with_name(template_path.stem + "_filled.xlsx")
    wb.save(filled)
    log.info("Template filled with %d shipment(s): %s", len(shipments), filled.name)
    return filled


def upload_ship_file(driver, file_path):
    """Phase 4c: attach the filled file, click Upload shipping info, read result."""
    if SHIP_UPLOAD_URL.split("?")[0] not in driver.current_url:
        driver.get(SHIP_UPLOAD_URL)
        time.sleep(5)

    file_input = None
    deadline = time.time() + 30
    while time.time() < deadline and file_input is None:
        inputs = driver.find_elements(By.XPATH, "//input[@type='file']")
        file_input = inputs[0] if inputs else None
        if file_input is None:
            time.sleep(2)
    if file_input is None:
        screenshot(driver, "no_file_input")
        raise RuntimeError("File input for the upload dropzone not found")

    file_input.send_keys(str(file_path))
    log.info("File attached: %s", file_path.name)
    time.sleep(3)

    upload_btn = wait_for(
        driver,
        "//button[.//text()[normalize-space()='Upload shipping info']][not(@disabled)]",
        30, "enabled 'Upload shipping info' button",
    )
    js_click(driver, upload_btn)
    log.info("Clicked Upload shipping info — waiting for the review page.")

    # TikTok navigates to an "Add tracking info" review page that lists the
    # orders identified in the file (already-shipped / unknown orders are
    # filtered out). A file-format problem surfaces as an error toast instead.
    review = None
    deadline = time.time() + 60
    while time.time() < deadline and review is None:
        time.sleep(3)
        review = find_visible(driver, "//*[normalize-space(text())='Add tracking info']")
        err = find_visible(
            driver, "//*[contains(translate(text(), 'EF', 'ef'), 'error') or "
                    "contains(translate(text(), 'EF', 'ef'), 'failed')]")
        if err:
            screenshot(driver, "upload_error")
            verdict = err.text.strip()
            log.info("UPLOAD RESULT (FAILED): %s", verdict)
            return False, verdict
    if review is None:
        screenshot(driver, "upload_no_review_page")
        log.warning("UPLOAD RESULT: review page never appeared")
        return False, "review page never appeared"

    time.sleep(3)
    body = driver.find_element(By.TAG_NAME, "body").text
    if "No orders identified" in body:
        screenshot(driver, "upload_result")
        log.info("UPLOAD RESULT (REJECTED): file parsed OK but no eligible "
                 "orders identified (already shipped / not awaiting shipment).")
        return False, "No orders identified"

    # Orders were identified — confirm with the final submit button.
    rows = driver.find_elements(
        By.XPATH, "//table//tr[td]")
    log.info("Review page lists %d order row(s) — confirming submission.", len(rows))
    # The confirm button text is dynamic, e.g. "Submit 1 parcel" /
    # "Submit 3 parcels", so match on a contained keyword rather than the
    # exact label.
    submit = None
    for keyword in ("Submit", "Confirm", "Save", "Add tracking info", "Upload"):
        submit = find_visible(
            driver,
            f"//button[contains(normalize-space(.), '{keyword}')][not(@disabled)]")
        if submit:
            break
    if not submit:
        screenshot(driver, "upload_no_submit_btn")
        log.warning("UPLOAD RESULT: orders identified but no submit button found")
        return False, "no submit button on review page"
    # Read the label before clicking — the click navigates/re-renders the
    # page, making the element reference stale immediately afterwards.
    submit_label = (submit.text or "").strip() or "Submit"
    js_click(driver, submit)
    log.info("Clicked final '%s' — waiting for confirmation.", submit_label)

    # After the click TikTok shows a "Processing... please wait" screen, then
    # either a success state or navigates back to Manage orders. Wait out the
    # processing screen; success = explicit success text OR the review page
    # ("Add tracking info") having gone away with no error shown.
    saw_processing = False
    deadline = time.time() + 120
    while time.time() < deadline:
        time.sleep(3)
        body = driver.find_element(By.TAG_NAME, "body").text.lower()
        if "please wait" in body and "processing" in body:
            saw_processing = True
            continue
        if "error" in body or "failed" in body:
            screenshot(driver, "upload_failed")
            log.info("UPLOAD RESULT (FAILED): error after final submit")
            return False, "error after final submit"
        if "success" in body or "uploaded" in body or "submitted" in body:
            screenshot(driver, "upload_success")
            log.info("UPLOAD RESULT (SUCCESS): tracking info submitted for %d order(s)",
                     len(rows))
            return True, f"submitted {len(rows)} order(s)"
        review_gone = not find_visible(
            driver, "//*[normalize-space(text())='Add tracking info']")
        if saw_processing and review_gone:
            screenshot(driver, "upload_success")
            log.info("UPLOAD RESULT (SUCCESS): processing finished, tracking "
                     "submitted for %d order(s)", len(rows))
            return True, f"submitted {len(rows)} order(s)"
    screenshot(driver, "upload_result")
    log.warning("UPLOAD RESULT: no confirmation detected after final submit")
    return False, "no confirmation after final submit"


def tracking_ui_test(order_id, tracking, carrier, headless, dry_run=True):
    """
    Exercise the Add tracking info flow for ONE order.

    Defaults to a dry run: it opens the order, walks the Combine orders dialog
    if it appears, fills the tracking ID and shipping provider, and stops with
    everything staged but the Submit button NOT clicked. Pass --submit to go
    through with it.
    """
    log.info("TEST MODE: Add tracking info for TikTok order %s (%s via %s)%s",
             order_id, tracking, carrier, "  [DRY RUN]" if dry_run else "  [WILL SUBMIT]")
    manager = BrowserManager(profile_dir=str(PROJECT_ROOT / "chrome_profile"))
    driver = manager.start_browser(headless=headless)
    try:
        ensure_logged_in(driver, headless)
        ok = submit_tracking_via_ui(driver, [{
            "order_id": str(order_id),
            "provider": tiktok_provider(carrier),
            "tracking": tracking,
        }], dry_run=dry_run)
        log.info("Tracking UI test %s", "passed" if ok else "FAILED")
        return ok
    except Exception:
        log.exception("Tracking UI test failed")
        screenshot(driver, "tracking_ui_test_failure")
        raise
    finally:
        manager.close()


def upload_test(headless):
    """DEPRECATED (template path). Test the shipping-upload mechanics with an
    already-shipped order.

    Uses the completed order 576907531018934637 with its REAL PostNL
    tracking — TikTok should reject it (already shipped), which safely
    proves download/fill/attach/upload/verdict end to end.
    """
    manager = BrowserManager(profile_dir=str(PROJECT_ROOT / "chrome_profile"))
    driver = manager.start_browser(headless=headless)
    try:
        ensure_logged_in(driver, headless)
        template = download_ship_template(driver)
        filled = fill_ship_template(template, [{
            "order_id": "576907531018934637",
            "provider": tiktok_provider("PostNL"),
            "tracking": "3SOSVJ0979112",
        }])
        ok, verdict = upload_ship_file(driver, filled)
        log.info("Upload test finished — possible in headless: YES. Verdict: %s", verdict)
    except Exception:
        log.exception("Upload test failed")
        screenshot(driver, "upload_test_failure")
        raise
    finally:
        manager.close()


def fulfilment_test(tiktok_id, headless):
    """Test mode: Shopify lookup + fulfilment lookup for one TikTok order ID."""
    from services.shopify import client as shopify

    log.info("TEST MODE: fulfilment lookup for TikTok order %s", tiktok_id)
    order = shopify.find_order_by_tiktok_id(tiktok_id)
    if not order:
        log.error("TikTok order %s not found in Shopify — cannot continue.", tiktok_id)
        sys.exit(1)
    log.info("Shopify match: %s (%s / %s)", order["name"],
             order["displayFinancialStatus"], order["displayFulfillmentStatus"])

    manager = BrowserManager(profile_dir=str(PROJECT_ROOT / "chrome_profile"))
    driver = manager.start_browser(headless=headless)
    try:
        run_fulfilment_phase(driver, headless,
                             [{"tiktok_id": tiktok_id, "shopify_order": order}])
        log.info("Fulfilment test completed successfully.")
    except Exception:
        log.exception("Fulfilment test failed")
        screenshot(driver, "fulfilment_test_failure")
        raise
    finally:
        manager.close()


def main():
    setup_logging()
    # headless via --headless flag or HEADLESS_MODE=1 (project convention);
    # default is visible so you can watch what's going on.
    headless = "--headless" in sys.argv or os.environ.get("HEADLESS_MODE") == "1"
    log.info("Browser mode: %s", "headless" if headless else "visible")

    # Test mode: --upload-test exercises the shipping-info upload mechanics
    # with an already-shipped order (safe — TikTok rejects duplicates).
    if "--upload-test" in sys.argv:
        upload_test(headless)
        return

    # Test mode: --tracking-test <order_id> <tracking> [carrier] drives the
    # Add tracking info flow for a single order. Dry run unless --submit is
    # also passed, so it can be verified without touching a real order.
    if "--tracking-test" in sys.argv:
        idx = sys.argv.index("--tracking-test")
        if idx + 2 >= len(sys.argv):
            print("Usage: python tiktok_toship_export_raw.py --tracking-test "
                  "<tiktok_order_id> <tracking_id> [carrier] [--submit]")
            sys.exit(2)
        carrier = "PostNL"
        if idx + 3 < len(sys.argv) and not sys.argv[idx + 3].startswith("--"):
            carrier = sys.argv[idx + 3]
        tracking_ui_test(sys.argv[idx + 1], sys.argv[idx + 2], carrier,
                         headless, dry_run="--submit" not in sys.argv)
        return

    # Test mode: --fulfilment-test <tiktok_order_id> skips the TikTok export
    # and runs only the Shopify + fulfilment lookups for one order.
    if "--fulfilment-test" in sys.argv:
        idx = sys.argv.index("--fulfilment-test")
        if idx + 1 >= len(sys.argv):
            print("Usage: python tiktok_toship_export_raw.py --fulfilment-test <tiktok_order_id>")
            sys.exit(2)
        fulfilment_test(sys.argv[idx + 1], headless)
        return

    manager = BrowserManager(profile_dir=str(PROJECT_ROOT / "chrome_profile"))
    driver = manager.start_browser(headless=headless)
    try:
        log_phase("PHASE 1/3: Export awaiting-shipment orders from TikTok")
        ensure_logged_in(driver, headless)
        log.info("On TikTok orders page.")
        previous_top = run_export(driver)
        report_name = wait_for_new_report(driver, previous_top)
        csv_path = download_report(driver, report_name)
        records = extract_orders(csv_path)
        found, _missing = match_orders_in_shopify(records)
        if found:
            results = run_fulfilment_phase(driver, headless, found)
            run_upload_phase(driver, results)
        else:
            log_phase("FINAL SUMMARY")
            log.info("No orders matched in Shopify — nothing to look up in fulfilment.")
        log.info("")
        log.info("Workflow completed successfully.")
    except Exception:
        log.exception("Workflow FAILED — see error above and screenshot in logs/")
        screenshot(driver, "failure")
        raise
    finally:
        manager.close()


if __name__ == "__main__":
    main()
